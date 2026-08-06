from __future__ import annotations

import csv
import io
from datetime import date
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from fpdf import FPDF
from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from core.serialization import json_value
from database import get_db
from dependencies import require_admin

router = APIRouter(prefix="/reportes", tags=["Reportes"])

REPORTS = {
    "accesses": {
        "filename": "accesos",
        "headers": ["ID", "Fecha", "Usuario", "Matrícula", "Movimiento", "Lector", "Resultado"],
        "query": """
            SELECT ar.id, ar.occurred_at,
                   COALESCE(u.full_name, 'Usuario desconocido') AS user_name,
                   COALESCE(u.enrollment, 'Sin identificar') AS enrollment,
                   ar.movement, d.name AS reader, ar.result
            FROM access_records ar
            LEFT JOIN users u ON u.id = ar.user_id
            JOIN devices d ON d.id = ar.device_id
            WHERE DATE(ar.occurred_at) BETWEEN :start AND :end
            ORDER BY ar.occurred_at DESC
        """,
    },
    "reservations": {
        "filename": "reservas",
        "headers": ["ID", "Fecha", "Inicio", "Fin", "Cubículo", "Usuario", "Matrícula", "Estado"],
        "query": """
            SELECT r.id, r.reservation_date, r.start_time, r.end_time,
                   c.name AS cubicle, u.full_name AS user_name,
                   u.enrollment, r.status
            FROM reservations r
            JOIN cubicles c ON c.id = r.cubicle_id
            JOIN users u ON u.id = r.user_id
            WHERE r.reservation_date BETWEEN :start AND :end
            ORDER BY r.reservation_date DESC, r.start_time DESC
        """,
    },
    "loans": {
        "filename": "prestamos",
        "headers": ["ID", "Usuario", "Matrícula", "Código", "Material", "Préstamo", "Vencimiento", "Devolución", "Estado"],
        "query": """
            SELECT l.id, u.full_name AS user_name, u.enrollment,
                   m.resource_code, m.title,
                   l.loan_date, l.due_date, l.return_date,
                   CASE
                     WHEN l.status IN ('active', 'renewed') AND l.due_date < CURDATE()
                     THEN 'overdue' ELSE l.status
                   END AS status
            FROM loans l
            JOIN users u ON u.id = l.user_id
            JOIN materials m ON m.id = l.material_id
            WHERE l.loan_date BETWEEN :start AND :end
            ORDER BY l.loan_date DESC
        """,
    },
    "users": {
        "filename": "usuarios",
        "headers": ["ID", "Nombre", "Correo", "Matrícula", "Rol", "Estado", "Registro"],
        "query": """
            SELECT u.id, u.full_name, u.email, u.enrollment,
                   r.name AS role_name, u.status, u.created_at
            FROM users u
            JOIN roles r ON r.id = u.role_id
            WHERE DATE(u.created_at) BETWEEN :start AND :end
            ORDER BY u.created_at DESC
        """,
    },
}


def _query_report(db: Session, report_type: str, start: date, end: date):
    definition = REPORTS.get(report_type)
    if definition is None:
        raise HTTPException(status_code=404, detail="El tipo de reporte no existe.")
    if start > end:
        raise HTTPException(status_code=422, detail="La fecha inicial no puede superar la final.")

    rows = db.execute(
        text(definition["query"]),
        {"start": start, "end": end},
    ).all()
    values = [[json_value(value) for value in row] for row in rows]
    return definition, values


def _csv_response(definition: dict, rows: list[list]):
    stream = io.StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow(definition["headers"])
    writer.writerows(rows)
    payload = "\ufeff" + stream.getvalue()
    return StreamingResponse(
        iter([payload]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename=Reporte_{definition['filename']}_SARA.csv"
        },
    )


def _excel_response(definition: dict, rows: list[list]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SARA"
    sheet.append(definition["headers"])
    for row in rows:
        sheet.append(row)
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 45)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=Reporte_{definition['filename']}_SARA.xlsx"
        },
    )


def _latin(value) -> str:
    return str(value if value is not None else "").encode("latin-1", "replace").decode("latin-1")


def _pdf_response(definition: dict, rows: list[list]):
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, _latin(f"Reporte S.A.R.A. - {definition['filename'].title()}"), ln=1)
    pdf.set_font("Arial", "B", 7)

    usable_width = 277
    column_width = usable_width / max(len(definition["headers"]), 1)
    for header in definition["headers"]:
        pdf.cell(column_width, 7, _latin(header)[:35], border=1)
    pdf.ln()

    pdf.set_font("Arial", size=6)
    for row in rows:
        for value in row:
            pdf.cell(column_width, 6, _latin(value)[:38], border=1)
        pdf.ln()

    payload = pdf.output(dest="S").encode("latin-1")
    return StreamingResponse(
        io.BytesIO(payload),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=Reporte_{definition['filename']}_SARA.pdf"
        },
    )


@router.get("/{report_type}/{report_format}")
def generate_report(
    report_type: Literal["accesses", "reservations", "loans", "users"],
    report_format: Literal["csv", "excel", "pdf"],
    inicio: date = Query(...),
    fin: date = Query(...),
    current_user: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    definition, rows = _query_report(db, report_type, inicio, fin)
    db.execute(
        text(
            """
            INSERT INTO audit_logs (
                actor_user_id, action, module, entity_type, entity_id, record_label
            ) VALUES (
                :actor, :action, 'Reportes', 'report', :entity_id, :record_label
            )
            """
        ),
        {
            "actor": current_user["id"],
            "action": f"Generó reporte en {report_format.upper()}",
            "entity_id": report_type,
            "record_label": f"{inicio.isoformat()} a {fin.isoformat()}",
        },
    )
    db.commit()
    if report_format == "csv":
        return _csv_response(definition, rows)
    if report_format == "excel":
        return _excel_response(definition, rows)
    return _pdf_response(definition, rows)

