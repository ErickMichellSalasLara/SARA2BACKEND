from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from database import get_db
import openpyxl
from fpdf import FPDF
import csv
import os

router = APIRouter()

# ---------------------------------------------------------
# 1. Exportar a CSV
# ---------------------------------------------------------
@router.get("/exportar/csv")
async def exportar_csv(db: Session = Depends(get_db)):
    try:
        query = text("SELECT * FROM vw_audit_records ORDER BY occurred_at DESC")
        registros = db.execute(query).mappings().all()

        file_path = "reporte_auditoria.csv"

        # Abrimos el archivo en modo escritura ('w')
        with open(file_path, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)

            # Escribimos los encabezados de las columnas
            writer.writerow(["ID", "Administrador", "Acción", "Módulo", "Registro", "Fecha"])

            # Escribimos fila por fila
            for reg in registros:
                writer.writerow([
                    reg["id"],
                    reg["administrator"],
                    reg["action"],
                    reg["module"],
                    reg["record_label"],
                    str(reg["occurred_at"])
                ])

        return FileResponse(file_path, filename="Reporte_SARA.csv")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ---------------------------------------------------------
# 2. Exportar a Excel
# ---------------------------------------------------------
@router.get("/exportar/excel")
async def exportar_excel(db: Session = Depends(get_db)):
    try:
        query = text("SELECT * FROM vw_audit_records ORDER BY occurred_at DESC")
        registros = db.execute(query).mappings().all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte SARA"

        ws.append(["ID", "Administrador", "Acción", "Módulo", "Registro", "Fecha"])

        for reg in registros:
            ws.append([
                reg["id"],
                reg["administrator"],
                reg["action"],
                reg["module"],
                reg["record_label"],
                str(reg["occurred_at"])
            ])

        file_path = "reporte_auditoria.xlsx"
        wb.save(file_path)
        return FileResponse(file_path, filename="Reporte_SARA.xlsx")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ---------------------------------------------------------
# 3. Exportar a PDF
# ---------------------------------------------------------
@router.get("/exportar/pdf")
async def exportar_pdf(db: Session = Depends(get_db)):
    try:
        # Para el PDF limitamos a 50 registros para que no sea un documento infinito
        query = text("SELECT * FROM vw_audit_records ORDER BY occurred_at DESC LIMIT 50")
        registros = db.execute(query).mappings().all()

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "S.A.R.A. - Reporte de Auditoría", ln=True, align="C")

        pdf.set_font("Arial", size=10)
        pdf.ln(10)

        for reg in registros:
            texto = f"[{reg['occurred_at']}] {reg['administrator']} -> {reg['action']} en {reg['module']}"
            pdf.cell(0, 10, texto, ln=True)

        file_path = "reporte_auditoria.pdf"
        pdf.output(file_path)
        return FileResponse(file_path, filename="Reporte_SARA.pdf")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))